import os
import openpyxl

from django.core.files.storage import FileSystemStorage

from data_handler.models import Activity


def file_save(file) -> str:
    """ Сохраняем файл """
    file_dir = os.getcwd() + "\\files\\"
    fs = FileSystemStorage(location=file_dir)
    path = fs.save(file.name, file)
    return file_dir + path


def file_read(path: str):
    """ Заносим считанные данные в базу """
    if path.__contains__('.txt'):
        with open(path, 'r') as file:
            for line in file:
                items = line.split(';')
                if len(items) > 0:
                    obj = Activity.objects.get_or_create(operation=items[0])
                    obj[0].N_count = items[1]
                    obj[0].n_error = items[2]
                    obj[0].k = items[3]
                    obj[0].Pk = items[4]
                    obj[0].Pob = items[5]
                    obj[0].Pi = items[6]
                    obj[0].save()
                else:
                    pass
    elif path.__contains__('.xlsx'):
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active
        for row_i in range(2, sheet_obj.max_row):
            obj = Activity.objects.get_or_create(operation=sheet_obj.cell(row=row_i, column=1).value)
            obj[0].N_count = sheet_obj.cell(row=row_i, column=3).value
            obj[0].n_error = sheet_obj.cell(row=row_i, column=4).value
            obj[0].k = sheet_obj.cell(row=row_i, column=5).value
            obj[0].Pk = sheet_obj.cell(row=row_i, column=7).value
            obj[0].Pob = sheet_obj.cell(row=row_i, column=8).value
            obj[0].Pi = sheet_obj.cell(row=row_i, column=9).value
            obj[0].save()
    return path
